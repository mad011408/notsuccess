"""NEXUS AI Agent - Excel Tool"""

from typing import Optional, List, Dict, Any
from dataclasses import dataclass, field


@dataclass
class ExcelSheet:
    """Excel sheet data"""
    name: str
    headers: List[str] = field(default_factory=list)
    rows: List[List[Any]] = field(default_factory=list)
    row_count: int = 0
    column_count: int = 0


@dataclass
class ExcelDocument:
    """Excel document"""
    path: str
    sheets: List[ExcelSheet] = field(default_factory=list)
    sheet_names: List[str] = field(default_factory=list)
    error: Optional[str] = None


class ExcelTool:
    """Read and write Excel files"""

    def read(
        self,
        path: str,
        sheet_name: Optional[str] = None,
        has_header: bool = True,
        limit: Optional[int] = None
    ) -> ExcelDocument:
        """
        Read Excel file

        Args:
            path: Excel file path
            sheet_name: Specific sheet to read (None = all)
            has_header: First row is header
            limit: Maximum rows to read

        Returns:
            ExcelDocument object
        """
        try:
            import openpyxl

            doc = ExcelDocument(path=path)
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            doc.sheet_names = wb.sheetnames

            sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames

            for name in sheets_to_read:
                if name not in wb.sheetnames:
                    continue

                ws = wb[name]
                sheet = ExcelSheet(name=name)

                rows_data = list(ws.iter_rows(values_only=True))

                if has_header and rows_data:
                    sheet.headers = [str(h) if h else '' for h in rows_data[0]]
                    rows_data = rows_data[1:]

                if limit:
                    rows_data = rows_data[:limit]

                sheet.rows = [list(row) for row in rows_data]
                sheet.row_count = len(sheet.rows)
                sheet.column_count = len(sheet.headers) if sheet.headers else (
                    len(sheet.rows[0]) if sheet.rows else 0
                )

                doc.sheets.append(sheet)

            wb.close()
            return doc

        except ImportError:
            return ExcelDocument(
                path=path,
                error="openpyxl not installed. Run: pip install openpyxl"
            )
        except Exception as e:
            return ExcelDocument(path=path, error=str(e))

    def read_as_dicts(
        self,
        path: str,
        sheet_name: Optional[str] = None,
        limit: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """Read sheet as list of dictionaries"""
        doc = self.read(path, sheet_name=sheet_name, has_header=True, limit=limit)
        if doc.error or not doc.sheets:
            return []

        sheet = doc.sheets[0]
        return [
            dict(zip(sheet.headers, row))
            for row in sheet.rows
        ]

    def write(
        self,
        path: str,
        data: List[List[Any]],
        sheet_name: str = "Sheet1",
        headers: Optional[List[str]] = None
    ) -> bool:
        """
        Write Excel file

        Args:
            path: Output path
            data: Row data
            sheet_name: Sheet name
            headers: Column headers

        Returns:
            Success status
        """
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            if headers:
                ws.append(headers)

            for row in data:
                ws.append(row)

            wb.save(path)
            return True

        except ImportError:
            return False
        except Exception:
            return False

    def write_dicts(
        self,
        path: str,
        data: List[Dict[str, Any]],
        sheet_name: str = "Sheet1",
        fieldnames: Optional[List[str]] = None
    ) -> bool:
        """Write list of dictionaries to Excel"""
        if not data:
            return False

        fieldnames = fieldnames or list(data[0].keys())
        rows = [[d.get(f) for f in fieldnames] for d in data]

        return self.write(path, rows, sheet_name, headers=fieldnames)

    def get_sheet_names(self, path: str) -> List[str]:
        """Get sheet names"""
        doc = self.read(path)
        return doc.sheet_names

    def get_sheet(
        self,
        path: str,
        sheet_name: str
    ) -> Optional[ExcelSheet]:
        """Get specific sheet"""
        doc = self.read(path, sheet_name=sheet_name)
        if doc.error or not doc.sheets:
            return None
        return doc.sheets[0]

    def to_csv(
        self,
        path: str,
        output_path: str,
        sheet_name: Optional[str] = None
    ) -> bool:
        """Convert Excel to CSV"""
        doc = self.read(path, sheet_name=sheet_name)
        if doc.error or not doc.sheets:
            return False

        import csv
        sheet = doc.sheets[0]

        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if sheet.headers:
                    writer.writerow(sheet.headers)
                writer.writerows(sheet.rows)
            return True
        except Exception:
            return False

    def get_cell(
        self,
        path: str,
        sheet_name: str,
        row: int,
        col: int
    ) -> Any:
        """Get specific cell value"""
        doc = self.read(path, sheet_name=sheet_name)
        if doc.error or not doc.sheets:
            return None

        sheet = doc.sheets[0]
        if row < len(sheet.rows) and col < len(sheet.rows[row]):
            return sheet.rows[row][col]
        return None

    def get_stats(self, path: str) -> Dict[str, Any]:
        """Get Excel file statistics"""
        doc = self.read(path)
        if doc.error:
            return {"error": doc.error}

        return {
            "sheets": len(doc.sheets),
            "sheet_names": doc.sheet_names,
            "sheet_details": [
                {
                    "name": s.name,
                    "rows": s.row_count,
                    "columns": s.column_count,
                    "headers": s.headers
                }
                for s in doc.sheets
            ]
        }

